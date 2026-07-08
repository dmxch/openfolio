"""Geteilte Excel-Reader fuer ETF-Holdings-Adapter.

SPDR/Fidelity liefern modernes .xlsx (OOXML) -> openpyxl.
HSBC liefert legacy .xls (BIFF/OLE2)          -> xlrd.
Beide geben Zeilen als list[list[str]] zurueck (Zellen zu getrimmten Strings
normalisiert), damit die Adapter-Parser rein string-basiert arbeiten koennen.
"""
from __future__ import annotations

import io
import logging

logger = logging.getLogger(__name__)


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # Ganzzahlige Floats ohne ".0" (Excel liefert oft 5.0 fuer 5).
        return str(int(v)) if v.is_integer() else repr(v)
    return str(v).strip()


def read_xlsx(content: bytes) -> list[list[str]]:
    """Erstes Worksheet eines .xlsx (OOXML) -> Liste von String-Zeilen. [] bei Fehler."""
    try:
        import openpyxl  # pure-Python, lazy import
    except ImportError:  # pragma: no cover
        logger.error("read_xlsx: openpyxl nicht installiert")
        return []
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
    except Exception as e:
        logger.warning("read_xlsx: load_workbook fehlgeschlagen: %s", e)
        return []
    try:
        ws = wb[wb.sheetnames[0]]
        return [[_cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    except Exception as e:
        logger.warning("read_xlsx: iter_rows fehlgeschlagen: %s", e)
        return []
    finally:
        try:
            wb.close()
        except Exception:
            pass


def read_xls(content: bytes) -> list[list[str]]:
    """Erstes Sheet eines legacy .xls (BIFF/OLE2) -> String-Zeilen. [] bei Fehler."""
    try:
        import xlrd  # pure-Python, lazy import
    except ImportError:  # pragma: no cover
        logger.error("read_xls: xlrd nicht installiert")
        return []
    try:
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        return [
            [_cell_to_str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
    except Exception as e:
        logger.warning("read_xls: open/parse fehlgeschlagen: %s", e)
        return []
